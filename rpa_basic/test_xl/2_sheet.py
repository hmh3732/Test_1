from openpyxl import Workbook
wb=Workbook()
ws=wb.create_sheet() #새로운 Sheet 생성
ws.title="Mysheet" #Sheet 이름 변경
ws.sheet_properties.tabColor="ff66ff"

ws1=wb.create_sheet("Yoursheet") #주어진 이름으로 Sheet 생성
ws1.sheet_properties.tabColor="b3ffcc"
ws2=wb.create_sheet("Newsheet")
wb.save("d:/2ndtype.xlsx")